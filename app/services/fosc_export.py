"""Export team time into FOSC weekly / quarterly attendance spreadsheet format.

The official contract process is quarterly; each quarter is a set of weekly
attendance sheets (Mon–Sun columns, In/Out/Phone/Offsite/FOSC/leave rows).
"""

from __future__ import annotations

from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.config import BASE_DIR
from app.models import (
    Employee, TimeEntry, OffsiteEntry, PhoneSupportEntry, DailySummary,
    EntryType, LeaveType, Role, TempoWeekly,
)

# Reference template (user may slim sheets; we generate a lean workbook)
TEMPLATE_FILENAME = (
    "2025-Q3 In Country - Weekly Time Record Worksheet Validation Template.xlsx"
)


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _excel_time(dt: datetime | None):
    if not dt:
        return None
    t = dt.time()
    return (t.hour * 3600 + t.minute * 60 + t.second) / 86400.0


def quarter_bounds(year: int, quarter: int) -> tuple[date, date]:
    """Return (first day of quarter, last day of quarter)."""
    if quarter not in (1, 2, 3, 4):
        raise ValueError("quarter must be 1–4")
    start_month = (quarter - 1) * 3 + 1
    start = date(year, start_month, 1)
    if quarter == 4:
        end = date(year, 12, 31)
    else:
        end = date(year, start_month + 3, 1) - timedelta(days=1)
    return start, end


def mondays_in_quarter(year: int, quarter: int) -> list[date]:
    """Every Monday that falls inside the quarter (weeks that touch the quarter)."""
    q_start, q_end = quarter_bounds(year, quarter)
    # Start from Monday of week containing first day
    m = _monday(q_start)
    mondays = []
    while m <= q_end:
        # Include week if any day Mon–Sun falls in quarter
        week_end = m + timedelta(days=6)
        if week_end >= q_start and m <= q_end:
            mondays.append(m)
        m += timedelta(days=7)
    return mondays


def _day_data(db: Session, emp_id: int, d: date) -> dict:
    entries = (
        db.query(TimeEntry)
        .filter(TimeEntry.employee_id == emp_id, TimeEntry.date == d)
        .order_by(TimeEntry.declared_time)
        .all()
    )
    cis = [e for e in entries if e.entry_type == EntryType.check_in]
    cos = [e for e in entries if e.entry_type == EntryType.check_out]
    first_in = min((e.declared_time for e in cis), default=None)
    last_out = max((e.declared_time for e in cos), default=None)

    phones = (
        db.query(PhoneSupportEntry)
        .filter(PhoneSupportEntry.employee_id == emp_id, PhoneSupportEntry.date == d)
        .all()
    )
    phone_h = round(sum(p.hours or 0 for p in phones), 2) or None

    offsites = (
        db.query(OffsiteEntry)
        .filter(OffsiteEntry.employee_id == emp_id, OffsiteEntry.date == d)
        .all()
    )
    off_h = 0.0
    for o in offsites:
        off_h += (o.end_time - o.start_time).total_seconds() / 3600.0
    off_h = round(off_h, 2) or None

    summary = (
        db.query(DailySummary)
        .filter(DailySummary.employee_id == emp_id, DailySummary.date == d)
        .first()
    )

    vac = sick = covid = hol = None
    fosc = None
    if summary:
        fosc = summary.total_hours if summary.total_hours else None
        if summary.leave_approved and summary.leave_hours:
            lt = summary.leave_type
            val = summary.leave_hours
            if lt == LeaveType.vacation:
                vac = val
            elif lt == LeaveType.sick:
                sick = val
            elif lt == LeaveType.covid_sick:
                covid = val
            elif lt == LeaveType.uae_holiday:
                hol = val

    return {
        "in": first_in,
        "out": last_out,
        "phone": phone_h,
        "offsite": off_h,
        "fosc": fosc,
        "vacation": vac,
        "sick": sick,
        "covid": covid,
        "holiday": hol,
    }


def _active_employees(db: Session) -> list[Employee]:
    employees = (
        db.query(Employee)
        .filter(Employee.is_active == True)  # noqa: E712
        .all()
    )
    return sorted(
        employees,
        key=lambda e: (0 if e.role == Role.employee else 1, e.name.lower()),
    )


def _style_constants():
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    name_font = Font(bold=True, size=11)
    cat_font = Font(size=10)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    return header_fill, header_font, name_font, cat_font, thin


def _tempo_hours(db: Session, emp_id: int, week_start: date) -> float | None:
    row = (
        db.query(TempoWeekly)
        .filter(
            TempoWeekly.employee_id == emp_id,
            TempoWeekly.week_start == week_start,
        )
        .first()
    )
    return float(row.hours) if row else None


def write_week_sheet(
    ws: Worksheet,
    db: Session,
    week_start: date,
    employees: list[Employee],
    *,
    title_suffix: str = "",
) -> list[dict]:
    """Fill one worksheet with a single week of FOSC attendance data.

    Returns per-employee metadata for the Discrepancy Tracker:
    [{employee, name, sdc_total, tempo_hours, totals_row, tempo_row}, ...]
    """
    week_start = _monday(week_start)
    days = [week_start + timedelta(days=i) for i in range(7)]
    header_fill, header_font, name_font, cat_font, thin = _style_constants()
    time_format = "HH:MM"
    num_format = "0.0"
    meta: list[dict] = []

    ws.merge_cells("A2:J2")
    title = "EADGE-T FOLLOW ON SERVICE CONTRACT - Weekly ATTENDANCE SHEET"
    if title_suffix:
        title = f"{title} ({title_suffix})"
    ws["A2"] = title
    ws["A2"].font = Font(bold=True, size=14, color="1F4E79")

    ws["A3"] = "Week Starting"
    ws["B3"] = datetime.combine(week_start, time.min)
    ws["B3"].number_format = "YYYY-MM-DD"
    ws["F3"] = "CONTRACT NO.: DP3/4/8/1/2021/7"

    # Row 4 headers match contract template (M = Hours to Reduce for INDEX col 13)
    headers = [
        "Name", "", "M", "T", "W", "T", "F", "S", "S",
        "Total Hours", "Signature", "Remarks", "Hours to Reduce",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, d in enumerate(days):
        cell = ws.cell(5, 3 + i, d)
        cell.number_format = "DD-MMM"
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=9, italic=True)

    row = 6
    for emp in employees:
        by_day = {d: _day_data(db, emp.id, d) for d in days}
        name_row = row

        ws.cell(row, 1, emp.name).font = name_font
        row += 1

        ws.cell(row, 1, "In").font = cat_font
        for i, d in enumerate(days):
            cell = ws.cell(row, 3 + i, _excel_time(by_day[d]["in"]))
            if cell.value is not None:
                cell.number_format = time_format
            cell.border = thin
        ws.cell(row, 12, "Time sheet should match TEMPO Inputs").font = Font(
            size=8, italic=True, color="666666"
        )
        row += 1

        ws.cell(row, 1, "Out").font = cat_font
        for i, d in enumerate(days):
            cell = ws.cell(row, 3 + i, _excel_time(by_day[d]["out"]))
            if cell.value is not None:
                cell.number_format = time_format
            cell.border = thin
        row += 1

        ws.cell(row, 1, "Phone Support Hours").font = cat_font
        for i, d in enumerate(days):
            cell = ws.cell(row, 3 + i, by_day[d]["phone"])
            if cell.value is not None:
                cell.number_format = num_format
            cell.border = thin
        cell_tot = ws.cell(row, 10, f"=SUM(C{row}:I{row})")
        cell_tot.number_format = num_format
        row += 1

        ws.cell(row, 1, "Off Site Work Hours").font = cat_font
        for i, d in enumerate(days):
            cell = ws.cell(row, 3 + i, by_day[d]["offsite"])
            if cell.value is not None:
                cell.number_format = num_format
            cell.border = thin
        cell_tot = ws.cell(row, 10, f"=SUM(C{row}:I{row})")
        cell_tot.number_format = num_format
        row += 1

        fosc_row = row
        ws.cell(row, 1, "Normal Time - FOSC").font = Font(bold=True, size=10)
        for i, d in enumerate(days):
            cell = ws.cell(row, 3 + i, by_day[d]["fosc"])
            if cell.value is not None:
                cell.number_format = num_format
            cell.border = thin
        cell_tot = ws.cell(row, 10, f"=SUM(C{row}:I{row})")
        cell_tot.number_format = num_format
        cell_tot.font = Font(bold=True)
        row += 1

        for label, key in [
            ("Vacation", "vacation"),
            ("Sick Hours", "sick"),
            ("COVID Sick Hours", "covid"),
            ("UAE National Holiday", "holiday"),
        ]:
            ws.cell(row, 1, label).font = cat_font
            for i, d in enumerate(days):
                cell = ws.cell(row, 3 + i, by_day[d][key])
                if cell.value is not None:
                    cell.number_format = num_format
                cell.border = thin
            cell_tot = ws.cell(row, 10, f"=SUM(C{row}:I{row})")
            cell_tot.number_format = num_format
            row += 1

        first_cat = fosc_row
        last_cat = row - 1
        totals_row = row
        first_name = emp.name.split()[0] if emp.name else "NAME"
        ws.cell(row, 1, f"Totals {first_name}").font = Font(bold=True, size=10)
        sdc_week = 0.0
        for i, d in enumerate(days):
            col = get_column_letter(3 + i)
            cell = ws.cell(row, 3 + i, f"=SUM({col}{first_cat}:{col}{last_cat})")
            cell.number_format = num_format
            cell.border = thin
            cell.font = Font(bold=True)
            day = by_day[d]
            day_total = 0.0
            for k in ("fosc", "vacation", "sick", "covid", "holiday"):
                if day[k]:
                    day_total += float(day[k])
            sdc_week += day_total
        # Template uses =SUM(J{fosc}:J{last_leave}) for week total on Totals row
        cell = ws.cell(row, 10, f"=SUM(J{first_cat}:J{last_cat})")
        cell.number_format = num_format
        cell.font = Font(bold=True)
        ws.cell(row, 11, "Signed: ")
        row += 1

        tempo_row = row
        tempo_h = _tempo_hours(db, emp.id, week_start)
        ws.cell(row, 1, "Hours from TEMPO").font = Font(size=9, italic=True, color="666666")
        for i in range(7):
            ws.cell(row, 3 + i).border = thin
        # Column J = week TEMPO total (Total Hours)
        tcell = ws.cell(row, 10, tempo_h if tempo_h is not None else None)
        if tempo_h is not None:
            tcell.number_format = num_format
        tcell.border = thin
        # L = label, M (col 13) = hours-to-reduce — same formula as contract template:
        #   =IF(SUM(J{totals}-J{tempo})>0, 0, SUM(J{totals}-J{tempo}))
        # Only shortfalls (base hours < TEMPO) are non-zero; excess is zeroed.
        ws.cell(row, 12, "Variance Reason:")
        vcell = ws.cell(
            row, 13,
            f"=IF(SUM(J{totals_row}-J{tempo_row})>0,0,SUM(J{totals_row}-J{tempo_row}))",
        )
        vcell.number_format = num_format
        vcell.border = thin

        # Discrepancy Tracker: INDEX(..., MATCH(name)+offset, 13)
        # offset = rows from employee name → Hours from TEMPO (always 11 with this block)
        name_to_tempo_offset = tempo_row - name_row
        row += 2

        meta.append({
            "employee_id": emp.id,
            "name": emp.name,
            "sdc_total": round(sdc_week, 2),
            "tempo_hours": tempo_h,
            "name_row": name_row,
            "totals_row": totals_row,
            "tempo_row": tempo_row,
            "name_to_tempo_offset": name_to_tempo_offset,
            "sheet_title": ws.title,
            "week_start": week_start,
        })

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 3
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 18
    ws.column_dimensions["M"].width = 14
    return meta


def _add_notes_sheet(wb: Workbook, lines: list[str]) -> None:
    leg = wb.create_sheet("Export Notes", 0)
    leg["A1"] = "FOSC Export — SDC Time Tracker"
    leg["A1"].font = Font(bold=True, size=12)
    for i, line in enumerate(lines, 2):
        leg[f"A{i}"] = line
    leg.column_dimensions["A"].width = 95


def _sheet_ref(title: str) -> str:
    """Excel sheet reference with quoting when needed."""
    if any(ch in title for ch in (" ", "'", "-", "!")):
        safe = title.replace("'", "''")
        return f"'{safe}'"
    return title


def write_discrepancy_sheet(
    ws: Worksheet,
    week_metas: list[tuple[str, date, list[dict]]],
    *,
    title: str = "Hours Discrepancies",
) -> None:
    """
    Discrepancy Tracker matching the contract workbook screenshot:

      A1  Hours Discrepancies
      A2  Note: only base hours less than TEMPO (obligation reductions)
      A3  No Data (status chip)
      B5… Wk1 … Wk14
      A6… employee names
      B6… INDEX/MATCH into each week sheet col M (Hours to Reduce)

    Exact formula pattern from their sheet (B6 example):
      =INDEX('Time Keeping Sheet (Wk1)'!$A$6:$M$999,
             MATCH('Discrepancy Tracker'!$A6,
                   'Time Keeping Sheet (Wk1)'!$A$6:$A$999,FALSE)+offset,13)

    offset = rows from employee name → Hours from TEMPO on that week sheet
    (11 with our full block: In/Out/Phone/Offsite/FOSC/4 leave/Totals/TEMPO).

    Col M on the week sheet:
      =IF(SUM(Jtotals-Jtempo)>0,0,SUM(Jtotals-Jtempo))
    """
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    peach = PatternFill("solid", fgColor="F8CBAD")
    total_fill = PatternFill("solid", fgColor="D9E2F3")
    num_format = "0.0"
    # Always show Wk1–Wk14 columns like the template, even if fewer weeks exist
    n_display = 14

    # ── Header block (screenshot layout) ─────────────────────────────
    ws["A1"] = title or "Hours Discrepancies"
    ws["A1"].font = Font(bold=True, size=16, color="000000")

    ws.merge_cells("A2:O2")
    ws["A2"] = (
        "Note: Only hours on base that are less than TEMPO hours are listed below, "
        "and are potential reductions to Quarterly Obligation Sheet. "
        "Any hours on base above TEMPO hours are not included."
    )
    ws["A2"].font = Font(size=9, italic=True, color="666666")

    ws["A3"] = "No Data"
    ws["A3"].fill = peach
    ws["A3"].font = Font(size=10, color="833C0C")

    # Row 5: Wk1 … Wk14
    for wi in range(n_display):
        cell = ws.cell(5, 2 + wi, f"Wk{wi + 1}")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    # Optional week-start dates on row 4 (not in screenshot; keep subtle for ops)
    for wi in range(min(n_display, len(week_metas))):
        _, week_start, _ = week_metas[wi]
        cell = ws.cell(4, 2 + wi, week_start)
        cell.number_format = "DD-MMM"
        cell.font = Font(size=8, color="888888")
        cell.alignment = Alignment(horizontal="center")

    employees_meta = week_metas[0][2] if week_metas else []
    default_offset = 11
    if employees_meta:
        default_offset = int(employees_meta[0].get("name_to_tempo_offset", 11))

    first_data_row = 6
    disc_ref = _sheet_ref(ws.title)  # usually 'Discrepancy Tracker'

    for ei, m0 in enumerate(employees_meta):
        r = first_data_row + ei
        ws.cell(r, 1, m0["name"]).font = Font(size=10)
        ws.cell(r, 1).border = thin

        for wi in range(n_display):
            cell = ws.cell(r, 2 + wi)
            cell.border = thin
            cell.number_format = num_format
            cell.alignment = Alignment(horizontal="center")

            if wi >= len(week_metas):
                # No week sheet for this column (e.g. short quarter / weekly export)
                cell.value = "-"
                continue

            sheet_title, _week_start, meta = week_metas[wi]
            offset = default_offset
            for m in meta:
                if m["name"] == m0["name"]:
                    offset = int(m.get("name_to_tempo_offset", default_offset))
                    break

            ref = _sheet_ref(sheet_title)
            # Match contract formula (screenshot formula bar):
            # INDEX(week!$A$6:$M$999, MATCH(disc!$A6, week!$A$6:$A$999, FALSE)+offset, 13)
            # IFERROR → "-" when name not found (same as #REF!/blank weeks in template)
            formula = (
                f"=IFERROR("
                f"INDEX({ref}!$A$6:$M$999,"
                f"MATCH({disc_ref}!$A{r},{ref}!$A$6:$A$999,FALSE)+{offset},13)"
                f",\"-\")"
            )
            cell.value = formula

    last_emp_row = (
        first_data_row + len(employees_meta) - 1
        if employees_meta
        else first_data_row
    )

    # Discrepancy Total (template row 34 style: =SUM(B6:B…))
    total_row = last_emp_row + 1 if employees_meta else first_data_row + 1
    ws.cell(total_row, 1, "Discrepancy Total").font = Font(bold=True, size=10)
    for wi in range(n_display):
        col = get_column_letter(2 + wi)
        cell = ws.cell(total_row, 2 + wi)
        if employees_meta and wi < len(week_metas):
            cell.value = f"=SUM({col}{first_data_row}:{col}{last_emp_row})"
        elif employees_meta:
            cell.value = 0
        else:
            cell.value = 0
        cell.font = Font(bold=True)
        cell.number_format = num_format
        cell.border = thin
        cell.fill = total_fill

    # Brief legend (below totals)
    note_row = total_row + 2
    ws.cell(note_row, 1, "How to read this sheet").font = Font(bold=True, size=11)
    ws.cell(note_row + 1, 1, (
        "Each cell uses INDEX/MATCH into that week’s Time Keeping sheet, column M "
        "(Hours to Reduce on the Hours from TEMPO row). "
        "Week formula: IF(SDC−TEMPO>0, 0, SDC−TEMPO) — only shortfalls (base < TEMPO). "
        "Import TEMPO weekly hours on Reports before exporting so column M is populated."
    ))
    ws.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 2, end_column=10)
    ws.cell(note_row + 1, 1).font = Font(size=9, color="555555")
    ws.cell(note_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 28
    for wi in range(n_display):
        ws.column_dimensions[get_column_letter(2 + wi)].width = 9


def _week_sheet_title(week_index: int) -> str:
    """Contract-style tab name: Time Keeping Sheet (Wk1) — max 31 chars."""
    return f"Time Keeping Sheet (Wk{week_index})"[:31]


def build_fosc_weekly_workbook(db: Session, week_start: date) -> BytesIO:
    """Single-week workbook (timekeeping + discrepancy tracker)."""
    week_start = _monday(week_start)
    employees = _active_employees(db)
    wb = Workbook()
    ws = wb.active
    ws.title = _week_sheet_title(1)
    meta = write_week_sheet(ws, db, week_start, employees)

    disc = wb.create_sheet("Discrepancy Tracker")
    write_discrepancy_sheet(
        disc,
        [(ws.title, week_start, meta)],
        title="Hours Discrepancies",
    )

    _add_notes_sheet(wb, [
        "",
        "Cadence: FOSC attendance packages are prepared each quarter.",
        f"This file: single week starting {_monday(week_start).isoformat()} (Monday).",
        "In / Out: declared check-in / check-out (first in, last out).",
        "Phone Support / Off Site: additive hours from the tracker.",
        "Normal Time - FOSC: clock + phone + offsite + BEOD credit.",
        "Leave rows: approved vacation / sick / COVID / UAE holiday hours.",
        "Hours from TEMPO: loaded from SDC when TEMPO weekly hours were imported.",
        "Discrepancy Tracker: INDEX/MATCH into each week sheet col M (Hours to Reduce).",
        "Only shortfalls (base hours < TEMPO) appear; excess is zeroed.",
        "",
        "Import TEMPO hours on Reports before exporting for full variance analysis.",
    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_fosc_quarterly_workbook(db: Session, year: int, quarter: int) -> BytesIO:
    """
    Full quarter package: one sheet per week + Discrepancy Tracker.

    Lean export (timekeeping only) — no multi-year Data dumps.
    Sheet names match the contract template: Time Keeping Sheet (Wk1) …
    """
    mondays = mondays_in_quarter(year, quarter)
    employees = _active_employees(db)
    wb = Workbook()
    default = wb.active
    week_metas: list[tuple[str, date, list[dict]]] = []

    for idx, mon in enumerate(mondays, 1):
        title = _week_sheet_title(idx)
        if idx == 1:
            ws = default
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        meta = write_week_sheet(
            ws, db, mon, employees,
            title_suffix=f"Q{quarter} Wk{idx}",
        )
        week_metas.append((ws.title, mon, meta))

    disc = wb.create_sheet("Discrepancy Tracker")
    write_discrepancy_sheet(
        disc,
        week_metas,
        title="Hours Discrepancies",
    )

    q_start, q_end = quarter_bounds(year, quarter)
    _add_notes_sheet(wb, [
        "",
        f"Quarterly package: {year} Q{quarter} ({q_start.isoformat()} → {q_end.isoformat()}).",
        f"Sheets: {len(mondays)} weekly attendance tabs + Discrepancy Tracker.",
        "Discrepancy Tracker uses INDEX/MATCH into each Time Keeping Sheet (WkN), col M.",
        "Hours to Reduce = only shortfalls where base hours < TEMPO.",
        "Import TEMPO weekly hours on Reports before export for complete analysis.",
        "",
        "Generated by SDC Time Tracker.",
    ])
    wb.move_sheet("Export Notes", offset=-len(wb.sheetnames) + 1)
    try:
        wb.move_sheet("Discrepancy Tracker", offset=-len(wb.sheetnames) + 2)
    except Exception:
        pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def template_path() -> Path:
    return BASE_DIR / TEMPLATE_FILENAME
