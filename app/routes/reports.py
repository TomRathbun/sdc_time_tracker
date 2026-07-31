"""Report and audit trail routes."""

from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import get_current_employee
from fastapi import Form, File, UploadFile

from app.models import (
    Employee, Role, TimeEntry, OffsiteEntry, DailySummary,
    AuditLog, LeaveRequest, LeaveStatus, TempoWeekly,
)
from app.services.time_calc import get_target_hours
from app.services.audit import log_action

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def _monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


@router.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db)):
    """Show reports page."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.name).all()
    week_start = _monday(date.today())
    today = date.today()
    current_quarter = (today.month - 1) // 3 + 1

    return templates.TemplateResponse("reports.html", {
        "request": request,
        "employee": employee,
        "employees": employees,
        "report_data": None,
        "fosc_week_start": week_start.isoformat(),
        "fosc_year": today.year,
        "fosc_quarter": current_quarter,
    })


@router.get("/reports/export/fosc-weekly")
async def export_fosc_weekly(
    request: Request,
    week: str = Query(None),
    db: Session = Depends(get_db),
):
    """Export one week of FOSC attendance (all active employees)."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    if week:
        try:
            week_start = _monday(date.fromisoformat(week))
        except ValueError:
            week_start = _monday(date.today())
    else:
        week_start = _monday(date.today())

    from app.services.fosc_export import build_fosc_weekly_workbook

    buf = build_fosc_weekly_workbook(db, week_start)
    filename = f"FOSC_Weekly_Attendance_{week_start.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/export/fosc-quarterly")
async def export_fosc_quarterly(
    request: Request,
    year: int = Query(None),
    quarter: int = Query(None),
    db: Session = Depends(get_db),
):
    """Export full quarter: one lean weekly sheet per week (no bulk history dumps)."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    today = date.today()
    if year is None:
        year = today.year
    if quarter is None:
        quarter = (today.month - 1) // 3 + 1
    if quarter < 1 or quarter > 4:
        quarter = 1

    from app.services.fosc_export import build_fosc_quarterly_workbook

    buf = build_fosc_quarterly_workbook(db, year, quarter)
    filename = f"FOSC_Q{quarter}_{year}_Attendance.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/tempo", response_class=HTMLResponse)
async def tempo_import_page(
    request: Request,
    week: str = Query(None),
    db: Session = Depends(get_db),
):
    """Manage TEMPO weekly hours for discrepancy comparison."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.name).all()
    if week:
        try:
            week_start = _monday(date.fromisoformat(week))
        except ValueError:
            week_start = _monday(date.today())
    else:
        week_start = _monday(date.today())
    rows = (
        db.query(TempoWeekly)
        .filter(TempoWeekly.week_start == week_start)
        .all()
    )
    by_emp = {r.employee_id: r for r in rows}

    return templates.TemplateResponse("tempo_import.html", {
        "request": request,
        "employee": employee,
        "employees": employees,
        "week_start": week_start,
        "by_emp": by_emp,
        "message": request.query_params.get("msg"),
        "error": request.query_params.get("err"),
    })


@router.post("/reports/tempo")
async def tempo_import_submit(
    request: Request,
    week: str = Form(...),
    db: Session = Depends(get_db),
):
    """Save TEMPO weekly hours from form fields tempo_{employee_id}."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    try:
        week_start = _monday(date.fromisoformat(week))
    except ValueError:
        return RedirectResponse(url="/reports/tempo?err=Invalid+week", status_code=303)

    form = await request.form()
    saved = 0
    for key, val in form.items():
        if not str(key).startswith("tempo_"):
            continue
        try:
            emp_id = int(str(key).replace("tempo_", ""))
        except ValueError:
            continue
        raw = str(val).strip()
        if raw == "":
            # clear existing
            existing = (
                db.query(TempoWeekly)
                .filter(TempoWeekly.employee_id == emp_id, TempoWeekly.week_start == week_start)
                .first()
            )
            if existing:
                db.delete(existing)
                saved += 1
            continue
        try:
            hours = float(raw)
        except ValueError:
            continue
        existing = (
            db.query(TempoWeekly)
            .filter(TempoWeekly.employee_id == emp_id, TempoWeekly.week_start == week_start)
            .first()
        )
        if existing:
            existing.hours = hours
            existing.updated_at = datetime.utcnow()
        else:
            db.add(TempoWeekly(
                employee_id=emp_id,
                week_start=week_start,
                hours=hours,
                updated_at=datetime.utcnow(),
            ))
        saved += 1

    db.commit()
    log_action(
        db, action="import_tempo", entity_type="TempoWeekly",
        entity_id=None, employee_id=employee.id,
        new_values={"week_start": str(week_start), "rows_touched": saved},
        ip_address=request.client.host if request.client else "",
    )
    return RedirectResponse(
        url=f"/reports/tempo?msg=Saved+{saved}+TEMPO+row(s)+for+week+{week_start.isoformat()}",
        status_code=303,
    )


@router.post("/reports/tempo/csv")
async def tempo_import_csv(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    CSV import: employee_name,week_start,hours
    week_start = any date in the week (normalized to Monday).
    """
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    import csv
    import io

    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    # Allow flexible headers
    saved = 0
    errors = 0
    emps = {e.name.strip().lower(): e for e in db.query(Employee).filter(Employee.is_active == True).all()}

    for row in reader:
        # normalize keys
        keys = {k.strip().lower().replace(" ", "_"): v for k, v in row.items() if k}
        name = (keys.get("employee_name") or keys.get("name") or keys.get("employee") or "").strip()
        week_raw = (keys.get("week_start") or keys.get("week") or keys.get("monday") or "").strip()
        hours_raw = (keys.get("hours") or keys.get("tempo_hours") or keys.get("tempo") or "").strip()
        if not name or not week_raw or not hours_raw:
            errors += 1
            continue
        emp = emps.get(name.lower())
        if not emp:
            # try partial match last name
            matches = [e for n, e in emps.items() if name.lower() in n or n in name.lower()]
            emp = matches[0] if len(matches) == 1 else None
        if not emp:
            errors += 1
            continue
        try:
            week_start = _monday(date.fromisoformat(week_raw[:10]))
            hours = float(hours_raw)
        except ValueError:
            errors += 1
            continue
        existing = (
            db.query(TempoWeekly)
            .filter(TempoWeekly.employee_id == emp.id, TempoWeekly.week_start == week_start)
            .first()
        )
        if existing:
            existing.hours = hours
            existing.updated_at = datetime.utcnow()
        else:
            db.add(TempoWeekly(
                employee_id=emp.id,
                week_start=week_start,
                hours=hours,
                updated_at=datetime.utcnow(),
            ))
        saved += 1

    db.commit()
    return RedirectResponse(
        url=f"/reports/tempo?msg=CSV+import:+{saved}+saved,+{errors}+skipped",
        status_code=303,
    )


@router.get("/reports/compliance", response_class=HTMLResponse)
async def compliance_report(
    request: Request,
    employee_id: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db),
):
    """Generate a compliance report."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.name).all()

    report_data = None
    if employee_id and employee_id.isdigit():
        # Default handling
        if not end_date:
            end_date = date.today().isoformat()
        if not start_date:
            # Default to 1 week ago if no start date provided
            start_dt = date.fromisoformat(end_date) - timedelta(days=7)
            start_date = start_dt.isoformat()

        s_date = date.fromisoformat(start_date)
        e_date = date.fromisoformat(end_date)
        target_emp = db.query(Employee).filter(Employee.id == int(employee_id)).first()

        if target_emp:
            # Get daily summaries for the period
            summaries = db.query(DailySummary).filter(
                DailySummary.employee_id == int(employee_id),
                DailySummary.date >= s_date,
                DailySummary.date <= e_date,
            ).order_by(DailySummary.date).all()

            # Get time entries for the period
            entries = db.query(TimeEntry).filter(
                TimeEntry.employee_id == int(employee_id),
                TimeEntry.date >= s_date,
                TimeEntry.date <= e_date,
            ).order_by(TimeEntry.date, TimeEntry.declared_time).all()

            # Get offsite entries
            offsite = db.query(OffsiteEntry).filter(
                OffsiteEntry.employee_id == int(employee_id),
                OffsiteEntry.date >= s_date,
                OffsiteEntry.date <= e_date,
            ).order_by(OffsiteEntry.date, OffsiteEntry.start_time).all()

            # Get leave days
            leaves = db.query(LeaveRequest).filter(
                LeaveRequest.employee_id == int(employee_id),
                LeaveRequest.status == LeaveStatus.approved,
                LeaveRequest.start_date <= e_date,
                LeaveRequest.end_date >= s_date,
            ).all()

            # Effective hours include approved leave (same rule as DailySummary compliance)
            def _effective(s):
                leave = s.leave_hours if (s.leave_approved and s.leave_hours) else 0.0
                return (s.total_hours or 0.0) + leave

            total_worked = sum(_effective(s) for s in summaries)
            total_target = sum(s.target_hours for s in summaries)
            deviation_days = [s for s in summaries if not s.is_compliant]

            report_data = {
                "target_employee": target_emp,
                "start_date": s_date,
                "end_date": e_date,
                "summaries": summaries,
                "entries": entries,
                "offsite": offsite,
                "leaves": leaves,
                "total_worked": round(total_worked, 2),
                "total_target": total_target,
                "deviation_days": deviation_days,
                "compliance_pct": round((total_worked / total_target * 100) if total_target > 0 else 0, 1),
            }

    return templates.TemplateResponse("reports.html", {
        "request": request,
        "employee": employee,
        "employees": employees,
        "report_data": report_data,
        "selected_employee_id": employee_id,
        "selected_start": start_date,
        "selected_end": end_date,
    })


@router.get("/reports/export/excel")
async def export_excel(
    request: Request,
    employee_id: int = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...),
    db: Session = Depends(get_db),
):
    """Export compliance report as Excel."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    s_date = date.fromisoformat(start_date)
    e_date = date.fromisoformat(end_date)
    target_emp = db.query(Employee).filter(Employee.id == employee_id).first()

    if not target_emp:
        return RedirectResponse(url="/reports", status_code=303)

    entries = db.query(TimeEntry).filter(
        TimeEntry.employee_id == employee_id,
        TimeEntry.date >= s_date,
        TimeEntry.date <= e_date,
    ).order_by(TimeEntry.date, TimeEntry.declared_time).all()

    summaries = db.query(DailySummary).filter(
        DailySummary.employee_id == employee_id,
        DailySummary.date >= s_date,
        DailySummary.date <= e_date,
    ).order_by(DailySummary.date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    # Header
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Compliance Report — {target_emp.name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Period: {s_date} to {e_date}"
    ws["A2"].font = Font(italic=True)

    # Column headers
    col_headers = ["Date", "Day", "Declared Time", "Submission Time", "Type", "Location", "Comments"]
    col_font = Font(bold=True, color="FFFFFF")
    col_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, h in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = col_font
        cell.fill = col_fill
        cell.border = thin_border

    # Data rows
    for row_idx, entry in enumerate(entries, 5):
        ws.cell(row=row_idx, column=1, value=str(entry.date))
        ws.cell(row=row_idx, column=2, value=entry.date.strftime("%A"))
        ws.cell(row=row_idx, column=3, value=entry.declared_time.strftime("%H:%M"))
        ws.cell(row=row_idx, column=4, value=entry.submission_time.strftime("%H:%M:%S"))
        ws.cell(row=row_idx, column=5, value=entry.entry_type.value.replace("_", " ").title())
        ws.cell(row=row_idx, column=6, value=entry.location_type.value.title())
        ws.cell(row=row_idx, column=7, value=entry.comments or "")
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Summary sheet
    ws2 = wb.create_sheet("Daily Summary")
    sum_headers = ["Date", "Day", "Worked (h)", "Target (h)", "Compliant"]
    for i, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = col_font
        cell.fill = col_fill
        cell.border = thin_border

    green_fill = PatternFill(start_color="c3e6cb", end_color="c3e6cb", fill_type="solid")
    red_fill = PatternFill(start_color="f5c6cb", end_color="f5c6cb", fill_type="solid")

    for row_idx, s in enumerate(summaries, 2):
        ws2.cell(row=row_idx, column=1, value=str(s.date))
        ws2.cell(row=row_idx, column=2, value=s.date.strftime("%A"))
        ws2.cell(row=row_idx, column=3, value=s.total_hours)
        ws2.cell(row=row_idx, column=4, value=s.target_hours)
        comp_cell = ws2.cell(row=row_idx, column=5, value="Yes" if s.is_compliant else "No")
        comp_cell.fill = green_fill if s.is_compliant else red_fill
        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).border = thin_border

    # Auto-width
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_sheet.column_dimensions[col_letter].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"compliance_{target_emp.name.replace(' ', '_')}_{s_date}_{e_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/audit", response_class=HTMLResponse)
async def audit_page(
    request: Request,
    employee_id: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db),
):
    """Show audit trail."""
    employee = get_current_employee(request, db)
    if not employee or employee.role not in (Role.manager, Role.supervisor):
        return RedirectResponse(url="/login", status_code=303)

    employees = db.query(Employee).order_by(Employee.name).all()

    query = db.query(AuditLog).order_by(AuditLog.timestamp.desc())

    if employee_id and employee_id.isdigit():
        query = query.filter(AuditLog.employee_id == int(employee_id))
    if start_date:
        s = datetime.fromisoformat(start_date)
        query = query.filter(AuditLog.timestamp >= s)
    if end_date:
        e = datetime.fromisoformat(end_date)
        query = query.filter(AuditLog.timestamp <= e + timedelta(days=1))

    logs = query.limit(200).all()

    return templates.TemplateResponse("audit.html", {
        "request": request,
        "employee": employee,
        "employees": employees,
        "logs": logs,
        "selected_employee_id": employee_id,
        "selected_start": start_date,
        "selected_end": end_date,
    })
